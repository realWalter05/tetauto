from openpyxl import load_workbook
from person import Person
from document import Doc


class Excel:


    def __init__(self, excel_path):
        self.doc = Doc()
        self.excel_path = excel_path
        self.sheet = self.get_sheet()
        self.iterate()
        self.save()


    def get_sheet(self):
        return load_workbook(self.excel_path).active


    def iterate(self):
        for row in list(self.sheet.iter_rows())[2:]:
            person = Person(row[0].value, row[1].value, row[2].value, row[3].value)
            self.doc.add_page(person)


    def save(self):
        print("Saving the document :)")
        self.doc.save_document()
